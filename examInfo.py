import os
import json
from openpyxl import Workbook
from openpyxl import load_workbook
import time

dirs = []
fileList = []
currentDir = os.getcwd() #获取当前目录的字符串
folder = currentDir+r"\folder" #拼接路径
out = currentDir +r"\out" #输出的路径
dirs = os.listdir(folder)
xlsxFiles = []
records = []  #记录写入的文件的路径


#创建一个答卷类
class AnswerSheet:
    # pass
    def __init__(self, danxuan =[""]*10, panduan =[""]*10, jianda = [""]*10):
        self.name = ""
        self.banji = ""
        self.xuehao = ""
        self.kemu = ""
        self.danxuan = danxuan
        self.panduan = panduan
        self.jianda = jianda

    def show(self):
        print(self.name + "," + self.banji + "," + self.xuehao + "," + self.kemu)
        print(self.danxuan)
        print(self.panduan)
        print(self.jianda)


#定义函数获取指定文件夹内的所有文件
allFileList = []
def get_all(dir):

    get_dir = os.listdir(dir)  # 遍历当前目录，获取文件列表
    # os.chdir('D:\PY\平时成绩\大数据2班')  # 如果不是，改到目标路径
    # get_dir = 'D:\PY\平时成绩\大数据2班'

    for i in get_dir:

        sub_dir = os.path.join(dir, i)  # 把第一步获取的文件加入路径
        if os.path.isdir(sub_dir):  # 如果当前仍然是文件夹，递归调用
            get_all(sub_dir)
        else:
            fileDir = sub_dir # 如果当前路径不是文件夹，则把文件名放入列表
            allFileList.append(fileDir)

    return allFileList



#获取指定目录(folder)和子目录里的所有文件的完整路径
fileList = get_all(folder)

#找出xlsx文件,把完整路径放入xlsxFiles列表中
for i in fileList:
    if ".xlsx" == os.path.splitext(i)[1]:
        xlsxFiles.append(i)



#读取xlsx文件内容到对象
def readXlsx(xlsxPath):
    # 读取xlsx文件
    wb = load_workbook(xlsxPath)  # 加载xlsx文件
    ws = wb[wb.sheetnames[0]]  # 获取sheet对象
    # print(list(ws.rows)[0][0].value)

    # 取xlsx文件里的值放在对象的属性中
    answer = AnswerSheet()  # 创建一个试卷对象
    answer.name = list(ws.rows)[1][0].value
    answer.banji = list(ws.rows)[1][1].value
    answer.xuehao = list(ws.rows)[1][2].value
    answer.kemu = list(ws.rows)[1][3].value

    # print(len(answer.danxuan))
    answer.danxuan[0] = list(ws.rows)[4][2].value
    answer.danxuan[1] = list(ws.rows)[5][2].value
    answer.danxuan[2] = list(ws.rows)[6][2].value
    answer.danxuan[3] = list(ws.rows)[7][2].value
    answer.danxuan[4] = list(ws.rows)[8][2].value
    answer.danxuan[5] = list(ws.rows)[9][2].value
    answer.danxuan[6] = list(ws.rows)[10][2].value
    answer.danxuan[7] = list(ws.rows)[11][2].value
    answer.danxuan[8] = list(ws.rows)[12][2].value
    answer.danxuan[9] = list(ws.rows)[13][2].value

    answer.panduan[0] = list(ws.rows)[14][2].value
    answer.panduan[1] = list(ws.rows)[15][2].value
    answer.panduan[2] = list(ws.rows)[16][2].value
    answer.panduan[3] = list(ws.rows)[17][2].value
    answer.panduan[4] = list(ws.rows)[18][2].value

    answer.jianda[0] = list(ws.rows)[19][2].value
    answer.jianda[1] = list(ws.rows)[20][2].value
    return answer


# ans =  readXlsx(xlsxFiles[0])  #读到对象里

#创建或加载excel
outPath = out + "\\统计.xlsx"

if os.path.exists(outPath):
    wb = load_workbook(outPath,False,False,False)
    # wb = load_workbook()
else:
    wb = Workbook()

ws = wb[wb.sheetnames[0]]

# 写表头
currentRow = ws.max_row+1
ws.cell(currentRow,1).value = "姓名"
ws.cell(currentRow,2).value = "班级"
ws.cell(currentRow,3).value = "学号"
ws.cell(currentRow,4).value = "科目"
ws.cell(currentRow,5).value = "单选1"
ws.cell(currentRow,6).value = "单选2"
ws.cell(currentRow,7).value = "单选3"
ws.cell(currentRow,8).value = "单选4"
ws.cell(currentRow,9).value = "单选5"
ws.cell(currentRow,10).value = "单选6"
ws.cell(currentRow,11).value = "单选7"
ws.cell(currentRow,12).value = "单选8"
ws.cell(currentRow,13).value = "单选9"
ws.cell(currentRow,14).value = "单选10"

ws.cell(currentRow,15).value = "判断1"
ws.cell(currentRow,16).value = "判断2"
ws.cell(currentRow,17).value = "判断3"
ws.cell(currentRow,18).value = "判断4"
ws.cell(currentRow,19).value = "判断5"

ws.cell(currentRow,20).value = "简答1"
ws.cell(currentRow,21).value = "简答2"

# print("最大列是:")
# print(ws.max_column)
# print(ws["b2"].value ==None)   #True

#写数据
# currentRow = [1]
def writeToXslx(answerObj,sheetObj):
    currentRow = sheetObj.max_row+1
    sheetObj.cell(currentRow, 1).value = answerObj.name
    sheetObj.cell(currentRow, 2).value = answerObj.banji
    sheetObj.cell(currentRow, 3).value = answerObj.xuehao
    sheetObj.cell(currentRow, 4).value = answerObj.kemu
    sheetObj.cell(currentRow, 5).value = answerObj.danxuan[0]
    sheetObj.cell(currentRow, 6).value = answerObj.danxuan[1]
    sheetObj.cell(currentRow, 7).value = answerObj.danxuan[2]
    sheetObj.cell(currentRow, 8).value = answerObj.danxuan[3]
    sheetObj.cell(currentRow, 9).value = answerObj.danxuan[4]
    sheetObj.cell(currentRow, 10).value = answerObj.danxuan[5]
    sheetObj.cell(currentRow, 11).value = answerObj.danxuan[6]
    sheetObj.cell(currentRow, 12).value = answerObj.danxuan[7]
    sheetObj.cell(currentRow, 13).value = answerObj.danxuan[8]
    sheetObj.cell(currentRow, 14).value = answerObj.danxuan[9]
    sheetObj.cell(currentRow, 15).value = answerObj.panduan[0]
    sheetObj.cell(currentRow, 16).value = answerObj.panduan[1]
    sheetObj.cell(currentRow, 17).value = answerObj.panduan[2]
    sheetObj.cell(currentRow, 18).value = answerObj.panduan[3]
    sheetObj.cell(currentRow, 19).value = answerObj.panduan[4]
    sheetObj.cell(currentRow, 20).value = answerObj.jianda[0]
    sheetObj.cell(currentRow, 21).value = answerObj.jianda[1]




# 读写所有xlsx文件的数据
for k in xlsxFiles:
    ans = readXlsx(k)  #读到对象
    writeToXslx(ans, ws)  #写到xslx
    records.append(k)
    # print(k)


# 保存文件
if os.path.exists(outPath):
    # print("创建文件失败:"+ outPath + ",方件已存在")
    # decision = input("是否删除文件?(Y/N)")
    # if decision == 'Y' or 'y':
    #     os.remove(outPath)
        wb.save(outPath)
        print("文件更新成功:%s,添加了%d条记录" % (outPath, len(records)))
else:
    wb.save(outPath)
    print("创建文件成功: %s , 添加了%d条记录" % (outPath, len(records)))
print("日志文件存放在:"+out+"\\log.txt")

#写入日志
if os.path.exists(out+"\\log.txt"):
    os.remove(out+"\\log.txt")
logFile = open(out+"\\log.txt", "a")
logFile.write("本次共写入了[%d]个文件,写入时间是:[%s] \n\n" %(len(records),time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) ))
for rec in records:
    logFile.write(rec+"\n")
logFile.close()











